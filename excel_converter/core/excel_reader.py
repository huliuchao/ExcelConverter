"""Excel读取器

负责读取Excel文件并解析标准三行格式的数据。
"""

from typing import List, Dict, Any, Optional
from dataclasses import dataclass
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from ..utils.error_handler import ConversionError


@dataclass
class ExcelField:
    """Excel字段信息"""
    name: str
    type: str
    column_index: int
    array_columns: Optional[List[int]] = None  # 数组字段的所有列索引（用于横向展开）


@dataclass
class ExcelData:
    """Excel数据容器"""
    file_path: Path
    sheet_name: str
    fields: List[ExcelField]
    rows: List[Dict[str, Any]]


class ExcelReader:
    """Excel文件读取器"""
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls']
    
    def read_excel(self, file_path: Path, sheet_name: str) -> ExcelData:
        """读取Excel文件"""
        if not file_path.exists():
            raise ConversionError(f"Excel file not found: {file_path}")
        
        if file_path.suffix.lower() not in self.supported_formats:
            raise ConversionError(f"Unsupported file format: {file_path.suffix}")
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise ConversionError(f"Failed to load Excel file {file_path}: {e}")
        
        if sheet_name not in workbook.sheetnames:
            available_sheets = ', '.join(workbook.sheetnames)
            raise ConversionError(f"Sheet '{sheet_name}' not found in {file_path}. Available sheets: {available_sheets}")
        
        worksheet = workbook[sheet_name]
        
        fields = self._parse_header_rows(worksheet)
        rows = self._parse_data_rows(worksheet, fields)
        
        workbook.close()
        
        return ExcelData(
            file_path=file_path,
            sheet_name=sheet_name,
            fields=fields,
            rows=rows
        )
    
    def _parse_header_rows(self, sheet: Worksheet) -> List[ExcelField]:
        """解析标准三行格式头部"""
        if sheet.max_row < 3:
            raise ConversionError("Excel sheet must have at least 3 header rows (description, field name, field type)")
        
        fields = []
        
        field_names_row = 2
        field_types_row = 3
        
        col_idx = 1
        while col_idx <= sheet.max_column:
            field_name_cell = sheet.cell(row=field_names_row, column=col_idx)
            field_type_cell = sheet.cell(row=field_types_row, column=col_idx)
            
            field_name = self._get_cell_value(field_name_cell)
            field_type = self._get_cell_value(field_type_cell)
            
            # 如果字段名或类型为空，跳过该列
            if not field_name or not field_type:
                col_idx += 1
                continue
            
            field_name = str(field_name).strip()
            field_type = str(field_type).strip()
            
            # 检查是否为数组类型，支持横向单元格展开
            if field_type.startswith('array<') and field_type.endswith('>'):
                array_columns = [col_idx]
                next_col = col_idx + 1
                
                while next_col <= sheet.max_column:
                    next_field_name = self._get_cell_value(sheet.cell(row=field_names_row, column=next_col))
                    next_field_type = self._get_cell_value(sheet.cell(row=field_types_row, column=next_col))
                    
                    if not next_field_name or str(next_field_name).strip() == '':
                        array_columns.append(next_col)
                        next_col += 1
                    else:
                        break
                
                fields.append(ExcelField(
                    name=field_name,
                    type=field_type,
                    column_index=col_idx,
                    array_columns=array_columns if len(array_columns) > 1 else None
                ))
                
                col_idx = next_col
            else:
                fields.append(ExcelField(
                    name=field_name,
                    type=field_type,
                    column_index=col_idx
                ))
                col_idx += 1
        
        if not fields:
            raise ConversionError("No valid fields found in Excel header rows")
        
        return fields
    
    def _parse_data_rows(self, sheet: Worksheet, fields: List[ExcelField], start_row: int = 4) -> List[Dict[str, Any]]:
        """解析数据行"""
        rows = []
        
        for row_idx in range(start_row, sheet.max_row + 1):
            row_data = {}
            has_data = False
            
            for field in fields:
                if field.array_columns:
                    array_values = []
                    for col_idx in field.array_columns:
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell_value = self._get_cell_value(cell)
                        
                        if cell_value is not None and str(cell_value).strip():
                            has_data = True
                            array_values.append(str(cell_value).strip())
                    
                    if array_values:
                        row_data[field.name] = '|'.join(array_values)
                    else:
                        row_data[field.name] = None
                else:
                    cell = sheet.cell(row=row_idx, column=field.column_index)
                    cell_value = self._get_cell_value(cell)
                    
                    if cell_value is not None and str(cell_value).strip():
                        has_data = True
                    
                    converted_value = self._convert_cell_value(cell_value, field.type)
                    row_data[field.name] = converted_value
            
            if has_data:
                rows.append(row_data)
        
        return rows
    
    def _get_cell_value(self, cell) -> Any:
        """获取单元格值"""
        if cell.value is None:
            return None
        
        if hasattr(cell, 'data_type') and cell.data_type == 'f':
            return cell.value
        
        return cell.value
    
    def _convert_cell_value(self, cell_value: Any, field_type: str) -> Any:
        """转换单元格值到指定类型"""
        if cell_value is None or str(cell_value).strip() == '':
            return None
        
        try:
            # 基本类型转换
            if field_type == 'int':
                if isinstance(cell_value, (int, float)):
                    return int(cell_value)
                return int(str(cell_value).strip())
            
            elif field_type == 'float':
                if isinstance(cell_value, (int, float)):
                    return float(cell_value)
                return float(str(cell_value).strip())
            
            elif field_type == 'bool':
                if isinstance(cell_value, bool):
                    return cell_value
                str_value = str(cell_value).strip().lower()
                return str_value in ['true', '1', 'yes', 'on', '是', '真']
            
            elif field_type == 'string':
                return str(cell_value).strip()
            
            else:
                # 对于复合类型（array:xxx, object:xxx），先返回字符串，后续由类型系统处理
                return str(cell_value).strip()
        
        except (ValueError, TypeError) as e:
            raise ConversionError(f"Failed to convert cell value '{cell_value}' to type '{field_type}': {e}")
    
    def validate_excel_format(self, file_path: Path, sheet_name: str) -> List[str]:
        """验证Excel格式是否符合标准三行格式"""
        errors = []
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            errors.append(f"Failed to load Excel file: {e}")
            return errors
        
        if sheet_name not in workbook.sheetnames:
            errors.append(f"Sheet '{sheet_name}' not found")
            workbook.close()
            return errors
        
        worksheet = workbook[sheet_name]
        
        if worksheet.max_row < 3:
            errors.append("Excel sheet must have at least 3 header rows")
        
        has_field_names = False
        has_field_types = False
        
        for col_idx in range(1, worksheet.max_column + 1):
            field_name = self._get_cell_value(worksheet.cell(row=2, column=col_idx))
            field_type = self._get_cell_value(worksheet.cell(row=3, column=col_idx))
            
            if field_name and str(field_name).strip():
                has_field_names = True
            
            if field_type and str(field_type).strip():
                has_field_types = True
        
        if not has_field_names:
            errors.append("Row 2 (field names) appears to be empty")
        
        if not has_field_types:
            errors.append("Row 3 (field types) appears to be empty")
        
        workbook.close()
        return errors
    
    def get_sheet_names(self, file_path: Path) -> List[str]:
        """获取Excel文件中的所有工作表名称"""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        except Exception as e:
            raise ConversionError(f"Failed to read sheet names from {file_path}: {e}")
    
    def get_field_preview(self, file_path: Path, sheet_name: str, max_rows: int = 5) -> Dict[str, Any]:
        """获取字段预览信息"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook[sheet_name]
            
            fields = self._parse_header_rows(worksheet)
            all_rows = self._parse_data_rows(worksheet, fields, start_row=4)
            preview_rows = all_rows[:max_rows]
            
            workbook.close()
            
            return {
                'fields': [{'name': f.name, 'type': f.type} for f in fields],
                'preview_rows': preview_rows,
                'total_rows': len(all_rows)
            }
        except Exception as e:
            raise ConversionError(f"Failed to get field preview: {e}")